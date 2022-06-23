
import openpyxl

excel_file = "data/jbfg_임직원.xlsx"

wb = openpyxl.load_workbook(excel_file)

# ws = wb.create_sheet('전북은행')
ws = wb['전북은행1']

# 데이터 추가 1
ws['A1'] = '사번'
ws['B1'] = '이름'
ws['C1'] = '부서명'
ws['D1'] = '연락처'
ws['E1'] = '이메일'

# 데이터 추가 2
# ws.cell(row=2,column=1,value=101)
# ws.cell(row=2,column=2,value='김은민')
# ws.cell(row=2,column=3,value='영업기획부')
# ws.cell(row=2,column=4,value='010-1234-1111')
# ws.cell(row=2,column=5,value='aa@kjbank.com')

# 데이터 추가 3

ws.append([201,'한혜형','디지털영업','010-1234-5678','aa@jbbank.com'])
ws.append([202,'김영목','카드사업부','010-1234-2222','bb@jbbank.com'])
ws.append([203,'박성실','데이터분석부','010-1234-3333','cc@jbbank.com'])
ws.append([204,'박요온','데이터분석부','010-1234-4444','dd@jbbank.com'])
ws.append([205,'오승현','지주','010-1234-5555','ee@jbbank.com'])
ws.append([206,'김진수','전략기획본부','010-1234-6666','ff@jbbank.com'])

# 시트명변경
# ws.title = '전북은행'

#워크시트 삭제
del wb['전북은행2']

wb.save(excel_file)
print('프로그램 종료')