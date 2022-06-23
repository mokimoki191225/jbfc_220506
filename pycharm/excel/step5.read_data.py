
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.utils import get_row

excel_file = "data/JBFG_임직원.xlsx"

wb = openpyxl.load_workbook(excel_file)
ws = wb.active

# 데이터조회

# A1 = ws['A1'].value
# print('A1 = {}' .format(A1))
#
# B1 = ws.cell(column=2,row=1).value
# print('B1 = {}' .format(B1))
#
# B2 = ws.cell(column=2,row=2).value
# print('B2 = {}' .format(B2))
#
# wb.close()
#
# print('프로그램 종료!!')


for b in range(1,6):
    for a in range(1,7):
        d ='{}{}'.format(get_column_letter(a),b)
        print(ws[d].value, end=' ')
    print() #한줄 띄는 방법#####

rows = ws.rows
print(rows)
print(type(rows))

arr_row = list()
cnt = 0

for row in rows:
    arr_row.append(row)
    # print(arr_row)
    for cell in row:
        val = cell.value
        print(val,end='\t')
    print()







