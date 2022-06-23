
import openpyxl

excel_file = "data/jbfg_임직원.xlsx"

wb = openpyxl.load_workbook(excel_file)

ws = wb.active

# 데이터 추가 1
ws['A1'] = '사번'
ws['B1'] = '이름'
ws['C1'] = '부서명'
ws['D1'] = '연락처'
ws['E1'] = '이메일'

# 데이터 추가 2
# ws.cell(row=2,column=1,value=101)
# ws.cell(row=2,column=2,value='김은민')
# ws.cell(row=2,column=3,value='영업기획부')
# ws.cell(row=2,column=4,value='010-1234-1111')
# ws.cell(row=2,column=5,value='aa@kjbank.com')

# 데이터 추가 3
ws.append([102,'박동현','IT기획부','010-1234-2222','bb@jbbank.com'])
ws.append([103,'이건호','수도권전략부','010-1234-3333','cc@jbbank.com'])
ws.append([104,'정은지','리스크관리부','010-1234-4444','dd@jbbank.com'])
ws.append([105,'최대훈','각화동지점','010-1234-5555','ee@jbbank.com'])

# 시트명변경
ws.title = '광주은행'

# 시트명추가

wb.save(excel_file)
print('프로그램 종료')