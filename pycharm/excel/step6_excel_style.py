
import openpyxl
from openpyxl.styles import PatternFill, colors, Font, Alignment

excel_file = "data/JBFG_임직원.xlsx"

wb = openpyxl.load_workbook(excel_file)
ws = wb.active

# 배경 색상
pattern_red = PatternFill(start_color='FF0000',fill_type='solid')# 빨간색 (R(FF)G(FF)B(FF))
pattern_yellow = PatternFill(start_color='FFFF00',fill_type='solid')
# pattern_red = PatternFill(start_color='0000FF',fill_type='solid')
pattern_blue = PatternFill(start_color=colors.BLUE,fill_type='solid')

ws.cell(1,2).fill = pattern_red
ws.cell(5,2).fill = pattern_yellow
ws.cell(4,2).fill = pattern_blue

# 폰트
font_20 = Font(name='나눔고딕', size=20, color=colors.BLUE)
ws.cell(1,1).font = font_20
ws.cell(2,2).font = font_20

# 정렬
align_center = Alignment(horizontal='center',vertical='center')

for idx in range(1,6):
    ws.cell(1, idx).alignment = align_center

for idx in range(1,7):
    ws.cell(idx, 2).alignment = align_center

# 셀크기
ws.row_dimensions[1].height = 50
ws.column_dimensions["C"].width = 15
ws.column_dimensions["D"].width = 20
ws.column_dimensions["E"].width = 25

wb.save(excel_file)
print('프로그램 종료!!')