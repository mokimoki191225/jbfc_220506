
import openpyxl
#컬럼 값을 숫자 값으로
from openpyxl.utils import get_column_letter

excel_file = "data/JBFG_임직원.xlsx"

wb = openpyxl.load_workbook(excel_file)
ws = wb['전북은행1']

# 데이터 삭제
# del ws['A7']

# row_no = 7
# col_nm = get_column_letter(2)
# print('{}{}'.format(col_nm, row_no))
#
# cell_loc = '{}{}'.format(col_nm, row_no)
# #value는 dict생각하면 된다.
# print('{} : {}' .format(cell_loc, ws[cell_loc].value))


# row_no = 7

# for col_no in range(1,6):
#     col_nm = get_column_letter(col_no)
#     print('{}{}' .format(col_nm,row_no))
#     # A7,B7,C7,D7
#     # cell_loc = '{}{}' .format(col_nm,row_no)
#     # del ws[cell_loc]

wb.save(excel_file)

print('프로그램 종료!!!')

# 다수의 행과 row 지우기
row_no = list(range(8,14))

for row_no in range(8,14):
    for col_no in range(1,6):
        col_nm = get_column_letter(col_no)
        # print('{}{}' .format(col_nm,row_no))
        # A7,B7,C7,D7 ~ A13,~E13
        cell_loc = '{}{}' .format(col_nm,row_no)
        del ws[cell_loc]

wb.save(excel_file)

print('프로그램 종료!!!')
