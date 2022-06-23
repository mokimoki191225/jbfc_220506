
import openpyxl

excel_file = "data/JBFG_임직원.xlsx"

wb = openpyxl.load_workbook(excel_file)
ws = wb['전북은행1']

#데이터 수정
ws['C7'] = '빅파이크래프트'
ws['D7'] = '010-1234-5568'
ws['E7'] = 'bigpycraft@gmail.com'

wb.save(excel_file)
print('프로그램 종료!!!')